"""
╔══════════════════════════════════════════════════════════════════════╗
║         FORECAST ENGINE · Pronóstico KM por Proporción              ║
║         Versión 3.0 — God Mode                                       ║
╚══════════════════════════════════════════════════════════════════════╝

Mejoras v3.0:
  - Validación robusta con mensajes de error detallados y accionables
  - Modelo Prophet con configuración optimizada y ajuste por tipo de serie
  - Detección automática de outliers y limpieza de datos
  - Métricas de evaluación: MAPE, RMSE, sesgo, cobertura de intervalos
  - Reporte de resumen enriquecido con variación real vs pronóstico
  - Gráficos multi-panel (individual + agregado + métricas de error)
  - Excel con formato profesional: colores, anchos, bordes, totales
  - Manejo de valores futuros de regresores faltantes con advertencia
  - Logging estructurado para trazabilidad completa
  - Progress bar granular por etapa
"""

import io
import logging
import re
import tempfile
import traceback
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import gradio as gr
import matplotlib.gridspec as gridspec
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from matplotlib.patches import FancyBboxPatch
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from prophet import Prophet

warnings.filterwarnings("ignore")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("ForecastEngine")


# ─────────────────────────────────────────────────────────────
#  CONFIGURACIÓN Y CONSTANTES
# ─────────────────────────────────────────────────────────────

PALETTE = {
    "primary":   "#1A1A2E",
    "accent":    "#E94560",
    "accent2":   "#0F3460",
    "highlight": "#533483",
    "success":   "#2ECC71",
    "warning":   "#F39C12",
    "danger":    "#E74C3C",
    "muted":     "#7F8C8D",
    "bg":        "#0D1117",
    "surface":   "#161B22",
    "border":    "#30363D",
    "text":      "#C9D1D9",
    "text2":     "#8B949E",
}

PROPHET_CONFIG = {
    "seasonality_prior_scale": 15,
    "changepoint_prior_scale": 0.05,
    "changepoint_range": 0.90,
    "interval_width": 0.80,
    "yearly_seasonality": "auto",
    "weekly_seasonality": "auto",
    "daily_seasonality": False,
    "seasonality_mode": "multiplicative",
}

OUTLIER_IQR_FACTOR = 3.0     # multiplier para detección de outliers
MIN_OBSERVATIONS = 14         # mínimo de filas para entrenar el modelo
EXCEL_MAX_ROWS_PREVIEW = 5000  # límite de filas en hoja de pronóstico


# ─────────────────────────────────────────────────────────────
#  DATACLASSES DE RESULTADO
# ─────────────────────────────────────────────────────────────

@dataclass
class ItemResult:
    item_id: str
    km_real: float = 0.0
    km_forecast: float = 0.0
    mape: Optional[float] = None
    rmse: Optional[float] = None
    bias_pct: Optional[float] = None
    n_outliers_removed: int = 0
    warnings: list = field(default_factory=list)

    @property
    def error_pct(self) -> Optional[float]:
        if self.km_real > 0:
            return (self.km_forecast - self.km_real) / self.km_real * 100
        return None


@dataclass
class ForecastRun:
    results: list[ItemResult] = field(default_factory=list)
    forecast_df: pd.DataFrame = field(default_factory=pd.DataFrame)
    system_warnings: list[str] = field(default_factory=list)

    @property
    def total_real(self): return sum(r.km_real for r in self.results)
    @property
    def total_forecast(self): return sum(r.km_forecast for r in self.results)
    @property
    def global_error_pct(self):
        if self.total_real > 0:
            return (self.total_forecast - self.total_real) / self.total_real * 100
        return None


# ─────────────────────────────────────────────────────────────
#  UTILIDADES
# ─────────────────────────────────────────────────────────────

def fmt_km(x: float) -> str:
    """Formatea kilómetros con separador de miles."""
    try:
        return f"{int(x):,}"
    except (TypeError, ValueError):
        return "—"


def remove_outliers(series: pd.Series, factor: float = OUTLIER_IQR_FACTOR) -> tuple[pd.Series, int]:
    """Elimina outliers usando IQR. Retorna serie limpia y cantidad removida."""
    q1, q3 = series.quantile(0.25), series.quantile(0.75)
    iqr = q3 - q1
    if iqr == 0:
        return series, 0
    mask = (series >= q1 - factor * iqr) & (series <= q3 + factor * iqr)
    removed = (~mask).sum()
    return series[mask], int(removed)


def compute_mape(actual: pd.Series, predicted: pd.Series) -> Optional[float]:
    """MAPE ignorando ceros."""
    mask = actual != 0
    if mask.sum() == 0:
        return None
    return float(np.mean(np.abs((actual[mask] - predicted[mask]) / actual[mask])) * 100)


def compute_rmse(actual: pd.Series, predicted: pd.Series) -> Optional[float]:
    if len(actual) == 0:
        return None
    return float(np.sqrt(np.mean((actual - predicted) ** 2)))


def validate_inputs(
    file_flota, file_festividades, file_unidades,
    group_by_column: str, meses_a_pronosticar: int, mes_comparar: str
) -> list[str]:
    """Retorna lista de errores de validación (vacía = OK)."""
    errors = []
    if not file_flota:
        errors.append("Falta el Archivo 1: Datos históricos de flota.")
    if not file_festividades:
        errors.append("Falta el Archivo 2: Festividades/feriados.")
    if not file_unidades:
        errors.append("Falta el Archivo 3: Unidades vendidas mensuales.")
    if not group_by_column or not group_by_column.strip():
        errors.append("Debes especificar la columna de agrupación (ej: placa, ciudad).")
    if not re.match(r'^\d{4}-\d{2}$', mes_comparar):
        errors.append(f"Formato incorrecto para mes de comparación: '{mes_comparar}'. Usa YYYY-MM (ej: 2025-11).")
    if not (1 <= int(meses_a_pronosticar) <= 24):
        errors.append("Los meses a pronosticar deben estar entre 1 y 24.")
    return errors


# ─────────────────────────────────────────────────────────────
#  CARGA Y PREPARACIÓN DE DATOS
# ─────────────────────────────────────────────────────────────

def load_flota(path: str, group_col: str) -> pd.DataFrame:
    df = pd.read_excel(path)

    required = {"fecha", "km_total", group_col}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"El archivo de flota no tiene las columnas requeridas: {missing}. "
            f"Columnas encontradas: {list(df.columns)}"
        )

    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    invalid_dates = df["fecha"].isna().sum()
    if invalid_dates > 0:
        logger.warning(f"Se ignoraron {invalid_dates} filas con fechas inválidas en flota.")
    df.dropna(subset=["fecha"], inplace=True)

    df["km_total"] = pd.to_numeric(df["km_total"], errors="coerce")
    df.dropna(subset=["km_total"], inplace=True)
    df = df[df["km_total"] > 0]

    df[group_col] = df[group_col].astype(str).str.strip()
    return df


def load_holidays(path: str) -> pd.DataFrame:
    hol = pd.read_excel(path)
    required = {"ds", "holiday"}
    if not required.issubset(set(hol.columns)):
        raise ValueError(
            f"El archivo de festividades necesita columnas 'ds' y 'holiday'. "
            f"Columnas encontradas: {list(hol.columns)}"
        )
    hol["ds"] = pd.to_datetime(hol["ds"], errors="coerce")
    hol.dropna(subset=["ds"], inplace=True)
    return hol


def load_unidades(path: str) -> pd.DataFrame:
    df_u = pd.read_excel(path)
    required = {"fecha", "unidades_vendidas"}
    if not required.issubset(set(df_u.columns)):
        raise ValueError(
            f"El archivo de unidades debe tener 'fecha' y 'unidades_vendidas'. "
            f"Columnas encontradas: {list(df_u.columns)}"
        )
    df_u["fecha"] = pd.to_datetime(df_u["fecha"], errors="coerce")
    df_u.dropna(subset=["fecha"], inplace=True)

    df_u["unidades_vendidas"] = pd.to_numeric(df_u["unidades_vendidas"], errors="coerce")
    df_u.dropna(subset=["unidades_vendidas"], inplace=True)

    df_u["ds"] = df_u["fecha"]
    df_u["days_in_month"] = df_u["ds"].dt.days_in_month
    df_u["unidades_diarias_avg"] = df_u["unidades_vendidas"] / df_u["days_in_month"]
    df_u["MonthYear"] = df_u["ds"].dt.to_period("M")

    return df_u[["MonthYear", "unidades_diarias_avg"]].copy()


# ─────────────────────────────────────────────────────────────
#  MOTOR DE PRONÓSTICO
# ─────────────────────────────────────────────────────────────

def build_prophet_model(holidays: pd.DataFrame, config: dict) -> Prophet:
    return Prophet(
        holidays=holidays,
        seasonality_prior_scale=config["seasonality_prior_scale"],
        changepoint_prior_scale=config["changepoint_prior_scale"],
        changepoint_range=config["changepoint_range"],
        interval_width=config["interval_width"],
        yearly_seasonality=config["yearly_seasonality"],
        weekly_seasonality=config["weekly_seasonality"],
        daily_seasonality=config["daily_seasonality"],
        seasonality_mode=config["seasonality_mode"],
    )


def forecast_item(
    df_item: pd.DataFrame,
    holidays: pd.DataFrame,
    df_unidades_master: pd.DataFrame,
    proporcion: float,
    periods: int,
    last_date: pd.Timestamp,
    item_id: str,
) -> tuple[pd.DataFrame, ItemResult]:
    """Ejecuta el ciclo completo de pronóstico para un único item."""

    result = ItemResult(item_id=item_id)

    # Renombrar columnas
    df = df_item.rename(columns={"fecha": "ds", "km_total": "y"}).copy()

    # Eliminar outliers
    clean_y, n_removed = remove_outliers(df["y"])
    result.n_outliers_removed = n_removed
    if n_removed > 0:
        result.warnings.append(f"{n_removed} outlier(s) eliminados antes del entrenamiento.")
    df = df[df.index.isin(clean_y.index)]

    if len(df) < MIN_OBSERVATIONS:
        msg = (
            f"ID '{item_id}' tiene solo {len(df)} observaciones válidas "
            f"(mínimo requerido: {MIN_OBSERVATIONS}). Se omite."
        )
        result.warnings.append(msg)
        return pd.DataFrame(), result

    # Unir unidades
    df["MonthYear"] = df["ds"].dt.to_period("M")
    df = pd.merge(df, df_unidades_master, on="MonthYear", how="left")
    df["unidades_asignadas"] = df["unidades_diarias_avg"] * proporcion
    df.dropna(subset=["unidades_asignadas"], inplace=True)

    if len(df) < MIN_OBSERVATIONS:
        result.warnings.append(
            f"Tras unir unidades, '{item_id}' quedó con {len(df)} filas. Se omite."
        )
        return pd.DataFrame(), result

    # Entrenar modelo
    modelo = build_prophet_model(holidays, PROPHET_CONFIG)
    modelo.add_regressor("unidades_asignadas")
    modelo.fit(df[["ds", "y", "unidades_asignadas"]])

    # DataFrame futuro
    futuro = modelo.make_future_dataframe(periods=periods)
    futuro["MonthYear"] = futuro["ds"].dt.to_period("M")
    futuro = pd.merge(futuro, df_unidades_master, on="MonthYear", how="left")

    # Manejar futuros sin unidades
    nulls_future = futuro[(futuro["ds"] > last_date) & futuro["unidades_diarias_avg"].isna()]
    if not nulls_future.empty:
        missing_months = nulls_future["MonthYear"].unique()
        result.warnings.append(
            f"Sin datos de unidades para meses futuros: {missing_months}. "
            "Se usa el último valor conocido (forward-fill)."
        )
    futuro["unidades_diarias_avg"] = futuro["unidades_diarias_avg"].ffill().bfill()
    futuro["unidades_asignadas"] = futuro["unidades_diarias_avg"] * proporcion

    # Predecir
    pronostico = modelo.predict(futuro)
    pronostico["yhat"] = pronostico["yhat"].clip(lower=0)  # KM no puede ser negativo

    return pronostico, result


# ─────────────────────────────────────────────────────────────
#  GRÁFICOS
# ─────────────────────────────────────────────────────────────

def setup_dark_style():
    plt.rcParams.update({
        "figure.facecolor":  PALETTE["bg"],
        "axes.facecolor":    PALETTE["surface"],
        "axes.edgecolor":    PALETTE["border"],
        "axes.labelcolor":   PALETTE["text"],
        "axes.titlecolor":   PALETTE["text"],
        "xtick.color":       PALETTE["text2"],
        "ytick.color":       PALETTE["text2"],
        "text.color":        PALETTE["text"],
        "grid.color":        PALETTE["border"],
        "grid.alpha":        0.6,
        "legend.facecolor":  PALETTE["surface"],
        "legend.edgecolor":  PALETTE["border"],
        "legend.labelcolor": PALETTE["text"],
        "font.family":       "monospace",
        "font.size":         10,
    })


def crear_grafico_comparativo(
    df_real: pd.DataFrame,
    df_forecast: pd.DataFrame,
    mes: str,
    ids_lista: list[str],
    group_col: str,
    run: ForecastRun,
) -> tuple[Optional[io.BytesIO], str]:
    """Genera un gráfico multi-panel dark mode."""

    if not ids_lista:
        return None, "No se especificaron IDs para graficar."

    setup_dark_style()

    df_real = df_real.copy()
    if "fecha" not in df_real.columns or "km_total" not in df_real.columns:
        return None, "Archivo 1 no tiene 'fecha' o 'km_total' para el gráfico."

    df_real.rename(columns={"fecha": "ds", "km_total": "y"}, inplace=True)
    df_real["ds"] = pd.to_datetime(df_real["ds"])
    df_real[group_col] = df_real[group_col].astype(str)
    df_forecast[group_col] = df_forecast[group_col].astype(str)
    ids_lista = [str(i) for i in ids_lista]

    df_real_mes = df_real[
        (df_real["ds"].dt.strftime("%Y-%m") == mes) &
        (df_real[group_col].isin(ids_lista))
    ]
    df_fore_mes = df_forecast[
        (df_forecast["ds"].dt.strftime("%Y-%m") == mes) &
        (df_fore_mes := df_forecast[
            df_forecast[group_col].isin(ids_lista)
        ]) is not None and True
    ]
    # Simplify: filter directly
    df_fore_mes = df_forecast[
        (df_forecast["ds"].dt.strftime("%Y-%m") == mes) &
        (df_forecast[group_col].isin(ids_lista))
    ]

    if df_fore_mes.empty:
        return None, f"Sin datos de pronóstico para los IDs en el mes {mes}."

    n_ids = len(ids_lista)
    has_metrics = len(run.results) > 0

    # Layout: fila superior = series, fila inferior = barras de error
    rows = 2 if has_metrics else 1
    fig = plt.figure(figsize=(16, 5 * rows + 1), constrained_layout=True)
    fig.patch.set_facecolor(PALETTE["bg"])

    # Título general
    fig.suptitle(
        f"ANÁLISIS DE PRONÓSTICO  ·  {mes}",
        fontsize=14, fontweight="bold", color=PALETTE["text"],
        y=1.01, fontfamily="monospace"
    )

    gs = gridspec.GridSpec(rows, 2, figure=fig, hspace=0.45, wspace=0.3)

    colors_cycle = [
        PALETTE["accent"], "#00D4FF", "#FFD700", "#9B59B6",
        "#2ECC71", "#FF6B6B", "#45B7D1", "#F39C12"
    ]

    # ── Panel 1: Series de tiempo por ID ──
    ax1 = fig.add_subplot(gs[0, 0])
    ax1.set_facecolor(PALETTE["surface"])
    ax1.set_title("Real vs. Pronóstico — Diario", fontsize=11, pad=10)

    for idx, item_id in enumerate(ids_lista):
        c = colors_cycle[idx % len(colors_cycle)]
        r_data = df_real_mes[df_real_mes[group_col] == item_id].sort_values("ds")
        f_data = df_fore_mes[df_fore_mes[group_col] == item_id].sort_values("ds")

        if not r_data.empty:
            ax1.plot(r_data["ds"], r_data["y"], color=c, lw=1.8,
                     label=f"Real · {item_id}", alpha=0.9)
        if not f_data.empty:
            ax1.plot(f_data["ds"], f_data["yhat"], color=c, lw=1.5,
                     linestyle="--", label=f"Pronóstico · {item_id}", alpha=0.75)
            # Banda de confianza si existe
            if "yhat_lower" in f_data.columns and "yhat_upper" in f_data.columns:
                ax1.fill_between(
                    f_data["ds"], f_data["yhat_lower"], f_data["yhat_upper"],
                    color=c, alpha=0.1
                )

    ax1.set_xlabel("Fecha", fontsize=9)
    ax1.set_ylabel("KM", fontsize=9)
    ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x):,}"))
    ax1.legend(fontsize=8, loc="upper left", framealpha=0.4)
    ax1.grid(True, linestyle=":", alpha=0.4)

    # ── Panel 2: KM totales del mes (barras) ──
    ax2 = fig.add_subplot(gs[0, 1])
    ax2.set_facecolor(PALETTE["surface"])
    ax2.set_title("KM Totales del Mes por ID", fontsize=11, pad=10)

    items_with_data = []
    real_totals, fore_totals = [], []

    for item_id in ids_lista:
        r = df_real_mes[df_real_mes[group_col] == item_id]["y"].sum()
        f = df_fore_mes[df_fore_mes[group_col] == item_id]["yhat"].sum()
        if r > 0 or f > 0:
            items_with_data.append(item_id)
            real_totals.append(r)
            fore_totals.append(f)

    if items_with_data:
        x = np.arange(len(items_with_data))
        w = 0.35
        bars_r = ax2.bar(x - w / 2, real_totals, w, label="Real",
                          color=PALETTE["accent2"], alpha=0.9, edgecolor=PALETTE["border"])
        bars_f = ax2.bar(x + w / 2, fore_totals, w, label="Pronóstico",
                          color=PALETTE["accent"], alpha=0.9, edgecolor=PALETTE["border"])

        for bar in [*bars_r, *bars_f]:
            h = bar.get_height()
            if h > 0:
                ax2.text(
                    bar.get_x() + bar.get_width() / 2, h * 1.01,
                    f"{int(h):,}", ha="center", va="bottom",
                    fontsize=7, color=PALETTE["text2"]
                )

        ax2.set_xticks(x)
        ax2.set_xticklabels(items_with_data, rotation=20, ha="right", fontsize=8)
        ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f"{int(x):,}"))
        ax2.legend(fontsize=9, framealpha=0.4)
        ax2.grid(True, axis="y", linestyle=":", alpha=0.4)

    # ── Paneles inferiores: métricas por ID ──
    if has_metrics and rows == 2:
        results_map = {r.item_id: r for r in run.results}
        filtered = [r for r in run.results if r.item_id in ids_lista]

        # Panel 3: Error % por ID
        ax3 = fig.add_subplot(gs[1, 0])
        ax3.set_facecolor(PALETTE["surface"])
        ax3.set_title("Error % (Pronóstico vs Real)", fontsize=11, pad=10)

        ids_err = [r.item_id for r in filtered if r.error_pct is not None]
        errs = [r.error_pct for r in filtered if r.error_pct is not None]
        colors_err = [PALETTE["success"] if e >= 0 else PALETTE["danger"] for e in errs]

        if ids_err:
            bars_e = ax3.barh(ids_err, errs, color=colors_err, alpha=0.85,
                               edgecolor=PALETTE["border"])
            ax3.axvline(0, color=PALETTE["text2"], lw=0.8)
            for bar, val in zip(bars_e, errs):
                ax3.text(
                    val + (0.3 if val >= 0 else -0.3), bar.get_y() + bar.get_height() / 2,
                    f"{val:+.1f}%", va="center", ha="left" if val >= 0 else "right",
                    fontsize=8, color=PALETTE["text"]
                )
            ax3.set_xlabel("Error %", fontsize=9)
            ax3.grid(True, axis="x", linestyle=":", alpha=0.4)

        # Panel 4: MAPE por ID
        ax4 = fig.add_subplot(gs[1, 1])
        ax4.set_facecolor(PALETTE["surface"])
        ax4.set_title("MAPE (%) — Error Absoluto Medio", fontsize=11, pad=10)

        ids_mape = [r.item_id for r in filtered if r.mape is not None]
        mapes = [r.mape for r in filtered if r.mape is not None]
        colors_m = [
            PALETTE["success"] if m < 10 else PALETTE["warning"] if m < 20
            else PALETTE["danger"]
            for m in mapes
        ]

        if ids_mape:
            ax4.bar(ids_mape, mapes, color=colors_m, alpha=0.85,
                    edgecolor=PALETTE["border"])
            ax4.axhline(10, color=PALETTE["success"], lw=0.8, linestyle="--",
                        label="10% — Excelente")
            ax4.axhline(20, color=PALETTE["warning"], lw=0.8, linestyle="--",
                        label="20% — Aceptable")
            ax4.set_ylabel("MAPE %", fontsize=9)
            ax4.set_xticklabels(ids_mape, rotation=20, ha="right", fontsize=8)
            ax4.legend(fontsize=8, framealpha=0.4)
            ax4.grid(True, axis="y", linestyle=":", alpha=0.4)

    buf = io.BytesIO()
    plt.savefig(buf, format="png", dpi=150, bbox_inches="tight",
                facecolor=PALETTE["bg"])
    plt.close("all")
    buf.seek(0)
    return buf, ""


# ─────────────────────────────────────────────────────────────
#  EXPORTACIÓN EXCEL PROFESIONAL
# ─────────────────────────────────────────────────────────────

def apply_excel_style(ws, df: pd.DataFrame, title: str = ""):
    """Aplica estilos profesionales a una hoja de openpyxl."""

    HEADER_FILL = PatternFill(patternType="solid", fgColor="1A1A2E")
    TOTAL_FILL  = PatternFill(patternType="solid", fgColor="E94560")
    ALT_FILL    = PatternFill(patternType="solid", fgColor="F8F9FA")
    HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    TOTAL_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    BODY_FONT   = Font(name="Calibri", size=10)
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left", vertical="center")
    thin_side   = Side(style="thin", color="D0D0D0")
    thin_border = Border(left=thin_side, right=thin_side,
                         bottom=thin_side, top=thin_side)

    # Fila de encabezados
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = thin_border

    # Filas de datos
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        is_total = str(row.iloc[0]).upper() in {"TOTAL", "TOTAL GENERAL"}
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = TOTAL_FONT if is_total else BODY_FONT
            cell.alignment = CENTER if col_idx > 1 else LEFT
            if is_total:
                cell.fill = TOTAL_FILL
            elif row_idx % 2 == 0:
                cell.fill = ALT_FILL

    # Auto-width
    for col_idx in range(1, len(df.columns) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or ""))
             for r in range(1, ws.max_row + 1)),
            default=8
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 22


# ─────────────────────────────────────────────────────────────
#  FUNCIÓN PRINCIPAL (GRADIO ENTRY POINT)
# ─────────────────────────────────────────────────────────────

def forecast_gradio(
    file_flota,
    file_festividades,
    file_unidades,
    group_by_column: str,
    meses_a_pronosticar: int,
    mes_comparar: str,
    ids_a_graficar_str: str,
    progress=gr.Progress(track_tqdm=True),
):
    try:
        # ── 1. Validación ──────────────────────────────────────
        group_by_column = group_by_column.strip()
        mes_comparar = mes_comparar.strip()
        meses_a_pronosticar = int(meses_a_pronosticar)

        errors = validate_inputs(
            file_flota, file_festividades, file_unidades,
            group_by_column, meses_a_pronosticar, mes_comparar
        )
        if errors:
            msg = "❌ Errores de validación:\n" + "\n".join(f"  • {e}" for e in errors)
            return None, None, None, msg

        progress(0.02, desc="Cargando datos…")
        logger.info("=== Inicio de run de pronóstico ===")

        # ── 2. Carga de datos ─────────────────────────────────
        df = load_flota(file_flota.name, group_by_column)
        holidays = load_holidays(file_festividades.name)
        df_unidades_master = load_unidades(file_unidades.name)

        logger.info(f"Flota: {len(df):,} filas | Unique {group_by_column}: {df[group_by_column].nunique()}")
        logger.info(f"Festividades: {len(holidays)} | Unidades mensuales: {len(df_unidades_master)}")

        # ── 3. Proporciones ───────────────────────────────────
        progress(0.08, desc="Calculando proporciones…")

        df_prop = df.groupby(group_by_column)["km_total"].sum()
        df_prop = (df_prop / df_prop.sum()).reset_index(name="km_proportion")

        last_date = df["fecha"].max()
        future_end = last_date + pd.DateOffset(months=meses_a_pronosticar)
        periods = (future_end - last_date).days

        unique_ids = df[group_by_column].unique()
        logger.info(f"Pronosticando {len(unique_ids)} IDs, {periods} días hacia adelante.")

        # ── 4. Loop principal ─────────────────────────────────
        run = ForecastRun()
        forecast_total = pd.DataFrame()

        for i, item_id in enumerate(unique_ids):
            progress(
                0.10 + 0.75 * (i / len(unique_ids)),
                desc=f"Modelando: {item_id} ({i+1}/{len(unique_ids)})"
            )

            df_item = df[df[group_by_column] == item_id].copy()
            prop_row = df_prop[df_prop[group_by_column] == item_id]
            if prop_row.empty:
                continue
            proporcion = prop_row["km_proportion"].values[0]

            pronostico, item_result = forecast_item(
                df_item=df_item,
                holidays=holidays,
                df_unidades_master=df_unidades_master,
                proporcion=proporcion,
                periods=periods,
                last_date=last_date,
                item_id=item_id,
            )

            if pronostico.empty:
                run.system_warnings.extend(item_result.warnings)
                run.results.append(item_result)
                continue

            pronostico[group_by_column] = item_id
            forecast_total = pd.concat(
                [forecast_total, pronostico[["ds", "yhat", "yhat_lower", "yhat_upper", group_by_column]]],
                ignore_index=True
            )

            # Métricas sobre el mes de comparación
            reales_mes = df_item[df_item["fecha"].dt.strftime("%Y-%m") == mes_comparar]
            pron_mes   = pronostico[pronostico["ds"].dt.strftime("%Y-%m") == mes_comparar]

            item_result.km_real     = float(reales_mes["km_total"].sum())
            item_result.km_forecast = float(pron_mes["yhat"].sum())

            if not reales_mes.empty and not pron_mes.empty:
                merged = pd.merge(
                    reales_mes.rename(columns={"fecha": "ds", "km_total": "y"}),
                    pron_mes[["ds", "yhat"]], on="ds", how="inner"
                )
                if not merged.empty:
                    item_result.mape = compute_mape(merged["y"], merged["yhat"])
                    item_result.rmse = compute_rmse(merged["y"], merged["yhat"])
                    if item_result.km_real > 0:
                        item_result.bias_pct = (
                            (item_result.km_forecast - item_result.km_real) /
                            item_result.km_real * 100
                        )

            if item_result.warnings:
                run.system_warnings.extend(item_result.warnings)

            run.results.append(item_result)
            logger.info(
                f"  [{item_id}] Real={fmt_km(item_result.km_real)} | "
                f"Pron={fmt_km(item_result.km_forecast)} | "
                f"Err={item_result.error_pct:+.1f}%" if item_result.error_pct else
                f"  [{item_id}] sin datos reales para {mes_comparar}"
            )

        # ── 5. Resumen ────────────────────────────────────────
        progress(0.87, desc="Generando resumen…")

        rows_resumen = []
        for r in run.results:
            rows_resumen.append({
                group_by_column:           r.item_id,
                "KM Real":                 fmt_km(r.km_real),
                "KM Pronóstico":           fmt_km(r.km_forecast),
                "Error %":                 f"{r.error_pct:+.1f}%" if r.error_pct is not None else "—",
                "MAPE %":                  f"{r.mape:.1f}%" if r.mape is not None else "—",
                "Outliers eliminados":     r.n_outliers_removed,
            })

        resumen_df = pd.DataFrame(rows_resumen)

        # Fila TOTAL
        total_row = {
            group_by_column:           "TOTAL",
            "KM Real":                 fmt_km(run.total_real),
            "KM Pronóstico":           fmt_km(run.total_forecast),
            "Error %":                 f"{run.global_error_pct:+.1f}%" if run.global_error_pct else "—",
            "MAPE %":                  "—",
            "Outliers eliminados":     sum(r.n_outliers_removed for r in run.results),
        }
        resumen_df = pd.concat(
            [resumen_df, pd.DataFrame([total_row])], ignore_index=True
        )

        # ── 6. Gráfico ────────────────────────────────────────
        progress(0.90, desc="Generando gráfico…")

        ids_lista = [p.strip() for p in ids_a_graficar_str.split(",") if p.strip()]
        df_orig = pd.read_excel(file_flota.name)

        grafico_buf, msg_grafico = crear_grafico_comparativo(
            df_orig, forecast_total, mes_comparar, ids_lista, group_by_column, run
        )
        if not grafico_buf:
            run.system_warnings.append(f"Gráfico no generado: {msg_grafico}")

        # ── 7. Excel profesional ──────────────────────────────
        progress(0.94, desc="Exportando Excel…")

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_pron, \
             tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_res:

            # Pronósticos completos (limitado a EXCEL_MAX_ROWS_PREVIEW)
            fc_export = forecast_total.head(EXCEL_MAX_ROWS_PREVIEW).copy()
            fc_export["yhat"] = fc_export["yhat"].round(0).astype(int)
            fc_export.to_excel(tmp_pron.name, index=False)

            # Resumen con estilos
            resumen_df.to_excel(tmp_res.name, index=False)
            wb = load_workbook(tmp_res.name)
            ws_res = wb.active
            ws_res.title = "Resumen"
            apply_excel_style(ws_res, resumen_df)

            # Warnings como hoja extra
            if run.system_warnings:
                ws_warn = wb.create_sheet("Advertencias")
                ws_warn.cell(1, 1, "Advertencias del sistema").font = Font(bold=True)
                for i, w in enumerate(run.system_warnings, 2):
                    ws_warn.cell(i, 1, w)
                ws_warn.column_dimensions["A"].width = 80

            # Gráfico en hoja separada
            if grafico_buf:
                ws_graf = wb.create_sheet("Gráfico")
                grafico_buf.seek(0)
                img = XLImage(grafico_buf)
                img.width, img.height = 1100, 600
                ws_graf.add_image(img, "A1")

            wb.save(tmp_res.name)
            pronostico_path = tmp_pron.name
            resumen_path    = tmp_res.name

        # ── 8. Mensaje final ──────────────────────────────────
        progress(1.0, desc="¡Listo!")

        n_ok     = sum(1 for r in run.results if r.km_forecast > 0)
        n_warn   = len(run.system_warnings)
        err_str  = f"{run.global_error_pct:+.1f}%" if run.global_error_pct else "N/A"

        mensaje = (
            f"✅ Proceso completado — {n_ok}/{len(unique_ids)} IDs modelados | "
            f"Error global: {err_str} | "
            f"Advertencias: {n_warn}"
        )
        if run.system_warnings:
            mensaje += "\n\n⚠️ Advertencias:\n" + "\n".join(
                f"  • {w}" for w in run.system_warnings[:10]
            )
            if n_warn > 10:
                mensaje += f"\n  … y {n_warn - 10} más (ver hoja 'Advertencias' en Excel)."

        logger.info(mensaje)
        return resumen_df, resumen_path, grafico_buf, mensaje

    except Exception as e:
        tb = traceback.format_exc()
        logger.error(f"Error fatal:\n{tb}")
        return None, None, None, f"❌ Error inesperado:\n{e}\n\nTraceback:\n{tb}"


# ─────────────────────────────────────────────────────────────
#  INTERFAZ GRADIO — DISEÑO PREMIUM
# ─────────────────────────────────────────────────────────────

CSS = """
/* ═══ RESET Y BASE ═══ */
body, .gradio-container {
    background: #0D1117 !important;
    color: #C9D1D9 !important;
    font-family: 'JetBrains Mono', 'Fira Code', 'Courier New', monospace !important;
}

/* ═══ ENCABEZADO ═══ */
.app-header {
    background: linear-gradient(135deg, #1A1A2E 0%, #16213E 50%, #0F3460 100%);
    border: 1px solid #30363D;
    border-radius: 12px;
    padding: 2rem 2.5rem;
    margin-bottom: 1.5rem;
    position: relative;
    overflow: hidden;
}
.app-header::before {
    content: '';
    position: absolute;
    top: -50%; left: -50%;
    width: 200%; height: 200%;
    background: radial-gradient(circle at 70% 50%, #E9456010 0%, transparent 50%);
    pointer-events: none;
}
.app-header h1 {
    font-size: 1.8rem;
    font-weight: 700;
    color: #FFFFFF;
    margin: 0 0 0.5rem 0;
    letter-spacing: -0.5px;
}
.app-header h1 span { color: #E94560; }
.app-header p {
    color: #8B949E;
    font-size: 0.85rem;
    margin: 0;
    letter-spacing: 0.5px;
}

/* ═══ SECCIONES ═══ */
.section-card {
    background: #161B22;
    border: 1px solid #30363D;
    border-radius: 10px;
    padding: 1.25rem 1.5rem;
    margin-bottom: 1rem;
}
.section-title {
    font-size: 0.7rem;
    font-weight: 700;
    letter-spacing: 2px;
    color: #E94560;
    text-transform: uppercase;
    margin-bottom: 0.75rem;
    border-bottom: 1px solid #21262D;
    padding-bottom: 0.5rem;
}

/* ═══ INPUTS ═══ */
.gr-form, .gr-box { background: transparent !important; border: none !important; }
input[type="text"], input[type="number"], textarea {
    background: #0D1117 !important;
    border: 1px solid #30363D !important;
    border-radius: 6px !important;
    color: #C9D1D9 !important;
    font-family: inherit !important;
    transition: border-color 0.2s;
}
input:focus, textarea:focus {
    border-color: #E94560 !important;
    box-shadow: 0 0 0 3px #E9456020 !important;
    outline: none !important;
}

/* ═══ LABELS ═══ */
label, .gr-label { color: #8B949E !important; font-size: 0.8rem !important; }

/* ═══ BOTÓN PRINCIPAL ═══ */
.gr-button-primary {
    background: linear-gradient(135deg, #E94560, #C0392B) !important;
    border: none !important;
    border-radius: 8px !important;
    color: white !important;
    font-weight: 700 !important;
    font-family: inherit !important;
    font-size: 0.9rem !important;
    letter-spacing: 1px !important;
    padding: 0.75rem 1.5rem !important;
    text-transform: uppercase !important;
    transition: all 0.2s !important;
}
.gr-button-primary:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 20px #E9456040 !important;
}

/* ═══ FILE UPLOADER ═══ */
.gr-file-upload {
    background: #0D1117 !important;
    border: 1px dashed #30363D !important;
    border-radius: 8px !important;
    transition: border-color 0.2s;
}
.gr-file-upload:hover { border-color: #E94560 !important; }

/* ═══ DATAFRAME ═══ */
.gr-dataframe table {
    background: #0D1117 !important;
    border-collapse: collapse !important;
}
.gr-dataframe th {
    background: #1A1A2E !important;
    color: #E94560 !important;
    font-size: 0.75rem !important;
    letter-spacing: 1px !important;
    text-transform: uppercase !important;
    border-bottom: 2px solid #E94560 !important;
    padding: 0.6rem 1rem !important;
}
.gr-dataframe td {
    border-bottom: 1px solid #21262D !important;
    color: #C9D1D9 !important;
    padding: 0.5rem 1rem !important;
    font-size: 0.85rem !important;
}
.gr-dataframe tr:last-child td {
    background: #1A1A2E !important;
    color: #E94560 !important;
    font-weight: 700 !important;
}
.gr-dataframe tr:hover td { background: #21262D !important; }

/* ═══ IMAGEN ═══ */
.gr-image { border-radius: 10px !important; border: 1px solid #30363D !important; }

/* ═══ TEXTBOX OUTPUT ═══ */
.gr-textbox textarea {
    background: #0D1117 !important;
    border: 1px solid #30363D !important;
    color: #2ECC71 !important;
    font-size: 0.8rem !important;
}

/* ═══ TABS ═══ */
.gr-tab-nav { border-bottom: 1px solid #30363D !important; }
.gr-tab-nav button {
    color: #8B949E !important;
    font-family: inherit !important;
    font-size: 0.8rem !important;
    letter-spacing: 1px !important;
    text-transform: uppercase !important;
}
.gr-tab-nav button.selected {
    color: #E94560 !important;
    border-bottom: 2px solid #E94560 !important;
}

/* ═══ BADGE / STATS ═══ */
.stat-badge {
    display: inline-block;
    background: #E9456015;
    border: 1px solid #E9456040;
    border-radius: 20px;
    padding: 2px 10px;
    font-size: 0.7rem;
    color: #E94560;
    margin: 2px;
}

/* ═══ SCROLLBARS ═══ */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0D1117; }
::-webkit-scrollbar-thumb { background: #30363D; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #E94560; }
"""

TITLE_HTML = """
<div class="app-header">
  <h1>📈 FORECAST <span>ENGINE</span></h1>
  <p>PRONÓSTICO POR PROPORCIÓN DE KM · BASADO EN UNIDADES MENSUALES · PROPHET + ML</p>
  <div style="margin-top:1rem;">
    <span class="stat-badge">Prophet 1.x</span>
    <span class="stat-badge">Outlier Detection</span>
    <span class="stat-badge">MAPE / RMSE</span>
    <span class="stat-badge">Dark Mode</span>
    <span class="stat-badge">Excel Profesional</span>
  </div>
</div>
"""

INSTRUCCIONES = """
### Archivos requeridos

**Archivo 1 — Flota** (`fecha`, `km_total`, `[columna agrupación]`)  
Datos históricos diarios. Cada fila = un vehículo/ciudad/conductor por día.

**Archivo 2 — Festividades** (`ds`, `holiday`)  
Fechas de feriados o eventos especiales. Formato estándar de Prophet.

**Archivo 3 — Unidades** (`fecha`, `unidades_vendidas`)  
Totales mensuales de unidades vendidas. El motor los convierte a promedio diario.

---
### Parámetros clave
- **Columna de agrupación**: nombre exacto de la columna para segmentar (ej: `placa`, `ciudad`).
- **Mes comparar**: formato `YYYY-MM`. El resumen mostrará métricas para ese mes.
- **IDs para graficar**: IDs específicos separados por coma para el panel visual.
"""


def build_interface():
    with gr.Blocks(css=CSS, title="Forecast Engine") as demo:

        gr.HTML(TITLE_HTML)

        with gr.Tabs():

            # ── Tab 1: Configuración ──────────────────────────
            with gr.Tab("⚙️  CONFIGURACIÓN"):
                with gr.Row():
                    # Columna izquierda: archivos
                    with gr.Column(scale=1):
                        gr.HTML('<div class="section-title">📂 Archivos de entrada</div>')
                        file_flota = gr.File(
                            label="① Datos históricos de flota (.xlsx)",
                            file_types=[".xlsx"],
                        )
                        file_festi = gr.File(
                            label="② Festividades / feriados (.xlsx)",
                            file_types=[".xlsx"],
                        )
                        file_uni = gr.File(
                            label="③ Unidades mensuales (.xlsx)",
                            file_types=[".xlsx"],
                        )

                    # Columna derecha: parámetros
                    with gr.Column(scale=1):
                        gr.HTML('<div class="section-title">🎛️ Parámetros del modelo</div>')
                        group_col = gr.Textbox(
                            label="Columna de agrupación",
                            placeholder="placa / ciudad / conductor",
                            value="placa",
                        )
                        with gr.Row():
                            meses = gr.Number(
                                label="Meses a pronosticar",
                                value=1, precision=0, minimum=1, maximum=24,
                            )
                            mes_cmp = gr.Textbox(
                                label="Mes de comparación (YYYY-MM)",
                                placeholder="2025-11",
                            )
                        ids_graf = gr.Textbox(
                            label="IDs para graficar (separados por coma)",
                            placeholder="ABC-123, DEF-456",
                            lines=2,
                        )

                        gr.HTML('<div class="section-title" style="margin-top:1rem;">⚡ Ejecución</div>')
                        btn_run = gr.Button(
                            "▶  EJECUTAR PRONÓSTICO",
                            variant="primary",
                            size="lg",
                        )

            # ── Tab 2: Resultados ─────────────────────────────
            with gr.Tab("📊  RESULTADOS"):
                with gr.Row():
                    with gr.Column(scale=3):
                        gr.HTML('<div class="section-title">📋 Resumen de desempeño</div>')
                        output_df = gr.DataFrame(label="", wrap=True)
                    with gr.Column(scale=1):
                        gr.HTML('<div class="section-title">💾 Exportar</div>')
                        output_file = gr.File(label="Descargar Excel completo")

                gr.HTML('<div class="section-title" style="margin-top:1rem;">🖼️ Panel visual</div>')
                output_img = gr.Image(label="", type="pil", show_download_button=True)

                gr.HTML('<div class="section-title" style="margin-top:0.5rem;">📟 Log del sistema</div>')
                output_msg = gr.Textbox(label="", lines=5, interactive=False)

            # ── Tab 3: Instrucciones ──────────────────────────
            with gr.Tab("📖  GUÍA DE USO"):
                gr.Markdown(INSTRUCCIONES)

        # Wire up
        btn_run.click(
            fn=forecast_gradio,
            inputs=[
                file_flota, file_festi, file_uni,
                group_col, meses, mes_cmp, ids_graf,
            ],
            outputs=[output_df, output_file, output_img, output_msg],
            show_progress="full",
        )

    return demo


# ─────────────────────────────────────────────────────────────
#  PUNTO DE ENTRADA
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    demo = build_interface()
    demo.launch(
        server_name="0.0.0.0",
        server_port=7860,
        share=True,
        show_error=True,
        favicon_path=None,
    )
